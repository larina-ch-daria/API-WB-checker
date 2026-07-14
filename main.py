"""
Everything is configured via env vars:
  WB_API_TOKEN     (required)  token with the "Statistics" scope
  WB_DATE_FROM     start of the pull, RFC3339/date    (default: 90 days ago)
  WB_STALE_DAYS    days-without-sales threshold        (default: 30)
  WB_OUTPUT        xlsx path                           (default: wb_supply_report.xlsx)
  WB_API_BASE      base host                           (default: https://statistics-api.wildberries.ru)
  WB_RATE_SLEEP    sleep between calls, sec (limit 1/min) (default: 60)
  WB_FETCH_STOCKS  pull Stocks for reconciliation (1/0) (default: 1)
"""
import os
import sys
import time
import logging
from datetime import datetime, timedelta, timezone

import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

log = logging.getLogger("wb")


# --- config
def cfg():
    token = os.environ.get("WB_API_TOKEN")
    if not token:
        sys.exit("WB_API_TOKEN is not set (needs the Statistics scope).")
    default_from = (datetime.now(timezone.utc) - timedelta(days=90)).strftime("%Y-%m-%d")
    return {
        "token": token,
        "date_from": os.environ.get("WB_DATE_FROM", default_from),
        "stale_days": int(os.environ.get("WB_STALE_DAYS", "30")),
        "output": os.environ.get("WB_OUTPUT", "wb_supply_report.xlsx"),
        "base": os.environ.get("WB_API_BASE", "https://statistics-api.wildberries.ru"),
        "rate_sleep": float(os.environ.get("WB_RATE_SLEEP", "60")),
        "fetch_stocks": os.environ.get("WB_FETCH_STOCKS", "1") not in ("0", "false", "False"),
    }


# ----fetch
def fetch_paginated(base, token, path, date_from, rate_sleep, extra=None):
    url = f"{base}{path}"
    headers = {"Authorization": token}
    cursor = date_from
    seen = set()
    rows, page = [], 0

    while True:
        params = {"dateFrom": cursor}
        if extra:
            params.update(extra)
        r = requests.get(url, headers=headers, params=params, timeout=60)
        if r.status_code == 429:  # rate limit, back off and retry the same cursor
            log.warning("429, sleeping %.0fs then retrying", rate_sleep)
            time.sleep(rate_sleep)
            continue
        r.raise_for_status()
        batch = r.json() or []
        page += 1
        log.info("%s: page %d, %d rows", path, page, len(batch))
        if not batch:
            break

        new = [x for x in batch if x.get("srid", x.get("saleID", id(x))) not in seen]
        for x in batch:
            seen.add(x.get("srid", x.get("saleID", id(x))))
        rows.extend(new)

        next_cursor = max(x.get("lastChangeDate", "") for x in batch)
        # bail if the cursor stalls or nothing new came back, otherwise this loops forever lol
        if next_cursor == cursor or not new:
            break
        cursor = next_cursor
        time.sleep(rate_sleep)

    return rows


def fetch_incomes(c):
    return fetch_paginated(c["base"], c["token"], "/api/v1/supplier/incomes",
                           c["date_from"], c["rate_sleep"])


def fetch_sales(c):
    return fetch_paginated(c["base"], c["token"], "/api/v1/supplier/sales",
                           c["date_from"], c["rate_sleep"], extra={"flag": 0})


def fetch_stocks(c):
    # Stocks is "current moment" only, no history kept; dateFrom is still required
    return fetch_paginated(c["base"], c["token"], "/api/v1/supplier/stocks",
                           c["date_from"], c["rate_sleep"])


# --- transforming
GROUP = ["income_id", "nm_id", "tech_size"]


def _to_date(s):
    if not s:
        return pd.NaT
    return pd.to_datetime(s, errors="coerce", utc=True).tz_localize(None)


def build_report(incomes, sales, stale_days):
    inc = pd.DataFrame(incomes).rename(columns={"incomeId": "income_id", "nmId": "nm_id",
                                                 "techSize": "tech_size"})
    sal = pd.DataFrame(sales)
    if inc.empty:
        raise SystemExit("Incomes came back empty — check the token and period.")

    inc["dateClose"] = inc["dateClose"].map(_to_date)
    received = (inc.groupby(GROUP, dropna=False)
                .agg(received=("quantity", "sum"),
                     close_date=("dateClose", "min"),          # a batch can close over several days
                     supplier_article=("supplierArticle", "first"),
                     warehouse=("warehouseName", "first"))
                .reset_index())

    diag = {"sales_rows": len(sal), "sales_income0": 0}
    if sal.empty:
        sold = pd.DataFrame(columns=GROUP + ["sold_net", "returns", "last_sale"])
    else:
        sal = sal.rename(columns={"incomeID": "income_id", "nmId": "nm_id", "techSize": "tech_size"})
        sal["date"] = sal["date"].map(_to_date)
        sal["kind"] = sal["saleID"].astype(str).str[0]   # saleID prefix: S=sale, R=return, D=surcharge
        diag["sales_income0"] = int((sal["income_id"].fillna(0) == 0).sum())  # unlinked sales, tracked for diag

        # 1 row = 1 unit and there's no reliable qty field here, so count rows
        s_sale = (sal[sal["kind"] == "S"].groupby(GROUP, dropna=False)
                  .agg(sold=("saleID", "size"), last_sale=("date", "max"))
                  .reset_index())
        s_ret = (sal[sal["kind"] == "R"].groupby(GROUP, dropna=False)
                 .agg(returns=("saleID", "size")).reset_index())
        sold = s_sale.merge(s_ret, on=GROUP, how="outer")
        sold["sold"] = sold["sold"].fillna(0).astype(int)
        sold["returns"] = sold["returns"].fillna(0).astype(int)
        sold["sold_net"] = sold["sold"] - sold["returns"]

    # left join off receipts: a sale whose income_id isn't among receipts (incl. 0) just drops
    df = received.merge(sold, on=GROUP, how="left")
    for col in ["sold", "returns", "sold_net"]:
        if col not in df:
            df[col] = 0
        df[col] = df[col].fillna(0).astype(int)

    # no per-batch stock in the API, deriving it
    df["stock_calc"] = df["received"] - df["sold_net"]

    today = pd.Timestamp(datetime.now().date())
    # if a batch had never been sold age it from the receipt date (not its missing last_sale)
    ref = df["last_sale"].fillna(df["close_date"]).dt.normalize()
    df["days_no_sales"] = (today - ref).dt.days.clip(lower=0)

    # avg over the batch's life
    span = (df["last_sale"] - df["close_date"]).dt.days
    df["velocity_per_day"] = (df["sold_net"] / span.where(span > 0)).round(2)

    def status(r):
        if r["stock_calc"] <= 0:
            return "Sold out"
        if r["days_no_sales"] > stale_days:
            return "Stale"
        return "Active"

    df["status"] = df.apply(status, axis=1)

    df = df.sort_values(["status", "stock_calc"], ascending=[True, False])
    cols = ["income_id", "close_date", "supplier_article", "nm_id", "tech_size", "warehouse",
            "received", "sold_net", "returns", "stock_calc",
            "last_sale", "days_no_sales", "velocity_per_day", "status"]
    return df[cols], diag


def build_reconcile(df, stocks):
    if not stocks:
        return None
    st = pd.DataFrame(stocks).rename(columns={"nmId": "nm_id"})
    st_g = st.groupby("nm_id")["quantityFull"].sum().rename("stock_reported")
    calc = df.groupby("nm_id")["stock_calc"].sum().rename("stock_derived")
    rec = pd.concat([calc, st_g], axis=1).fillna(0).reset_index()
    rec["gap"] = rec["stock_derived"] - rec["stock_reported"]
    return rec.sort_values("gap", key=lambda s: s.abs(), ascending=False)


# --- xlsx export
STATUS_FILL = {"Stale": "FFC7CE", "Active": "C6EFCE", "Sold out": "D9D9D9"}
HEAD_FILL = PatternFill("solid", fgColor="305496")


def _autofit(ws):
    for col in ws.columns:
        w = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max(w + 2, 10), 42)


def _write_sheet(ws, df, freeze=True):
    for j, h in enumerate(df.columns, 1):
        c = ws.cell(1, j, h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF")
        c.fill = HEAD_FILL
        c.alignment = Alignment(horizontal="center")
    for i, row in enumerate(df.itertuples(index=False), start=2):
        for j, v in enumerate(row, 1):
            if isinstance(v, pd.Timestamp):
                v = None if pd.isna(v) else v.date().isoformat()
            elif pd.isna(v):
                v = None
            ws.cell(i, j, v)
    if freeze:
        ws.freeze_panes = "A2"
    _autofit(ws)


def write_xlsx(df, diag, rec, path, stale_days):
    wb = Workbook()
    ws = wb.active
    ws.title = "Batches"
    _write_sheet(ws, df)
    status_col = list(df.columns).index("status") + 1  # tint the status cell per row
    for i, st in enumerate(df["status"], start=2):
        if st in STATUS_FILL:
            ws.cell(i, status_col).fill = PatternFill("solid", fgColor=STATUS_FILL[st])

    d = wb.create_sheet("Diagnostics")
    lines = [
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Stale threshold, days", stale_days),
        ("Sales rows total", diag.get("sales_rows", 0)),
        ("Sales without supply number (income_id=0)", diag.get("sales_income0", 0)),
        ("", ""),
        ("Note", "stock_calc = received - sold_net. The API has no per-batch stock."),
        ("Note", "sales keeps 90 days: for older batches pull sales from reportDetailByPeriod (gi_id)."),
    ]
    for i, (k, v) in enumerate(lines, 1):
        d.cell(i, 1, k).font = Font(name="Arial", bold=True)
        d.cell(i, 2, v)
    _autofit(d)

    if rec is not None and not rec.empty:
        _write_sheet(wb.create_sheet("Reconcile"), rec)

    wb.save(path)
    return path


# --- main
def main():
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
    c = cfg()
    log.info("From %s | stale threshold %d days", c["date_from"], c["stale_days"])
    incomes = fetch_incomes(c)
    sales = fetch_sales(c)
    stocks = fetch_stocks(c) if c["fetch_stocks"] else []
    df, diag = build_report(incomes, sales, c["stale_days"])
    rec = build_reconcile(df, stocks)
    out = write_xlsx(df, diag, rec, c["output"], c["stale_days"])
    log.info("Done: %s | %d batch rows | %d stale", out, len(df), int((df["status"] == "Stale").sum()))


if __name__ == "__main__":
    main()